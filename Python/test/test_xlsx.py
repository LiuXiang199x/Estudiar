import os
import sys
from openpyxl import Workbook


def demo():
    wb = Workbook() #创建工作簿
    ws = wb.active #激活工作表

    #可以通过指定单元格的形式来写入数据
    ws['A1'] = '姓名'
    ws['B1'] = '年龄'
    ws['C1'] = '班级'

    #也可以通过以下这种方式一次写入一行数据，列表中的第一个数据会写在A列，第二个数据会写在B列......
    ws.append(['张三',13,1])
    ws.append(['李四',12,2])
    ws.append(['王五',14,3])

    wb.save('stuInfo.xlsx') #保存文件

pic_path = "/home/agent/1.txt"
xlsx_path = "/home/agent/test.xlsx"
wb = Workbook()
ws = wb.active

ws.append(["=HYPERLINK(" + "\"" + pic_path + "\"" + ")", 0])

wb.save(xlsx_path)

